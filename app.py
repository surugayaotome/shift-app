import streamlit as st
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
from sqlalchemy.pool import NullPool
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import datetime

# --- AgGridのインポート ---
from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode, GridUpdateMode, JsCode

st.set_page_config(page_title="日本橋乙女 シフト管理", layout="wide")

# ==========================================
# 1. データベース接続＆超堅牢な初期化
# ==========================================
@st.cache_resource
def get_engine():
    try:
        raw_uri = st.secrets["database"]["uri"]
        prefix, rest = raw_uri.split("://")
        user_pass, host_db = rest.rsplit("@", 1)
        user, password = user_pass.split(":", 1)
        host_port, db_query = host_db.split("/", 1)
        host, port = host_port.split(":")
        db_name = db_query.split("?")[0]

        url_object = URL.create(
            drivername="postgresql+psycopg2",
            username=user,
            password=password,
            host=host,
            port=int(port),
            database=db_name,
            query={"sslmode": "require"},
        )
        return create_engine(url_object, poolclass=NullPool)
    except Exception as e:
        st.error(f"接続設定エラー: {e}")
        return None

engine = get_engine()

def init_db():
    if engine is None: return
    
    # 各命令ごとに個別のトランザクションで実行（道連れエラー防止）
    for cmd in [
        "CREATE TABLE IF NOT EXISTS shift_data (day TEXT, staff_name TEXT, off_status TEXT, shift_json TEXT, PRIMARY KEY (day, staff_name));",
        "CREATE TABLE IF NOT EXISTS staff_master (staff_name TEXT PRIMARY KEY, password TEXT NOT NULL, role_name TEXT, is_admin BOOLEAN DEFAULT FALSE);",
        "CREATE TABLE IF NOT EXISTS system_config (config_key TEXT PRIMARY KEY, config_value TEXT);"
    ]:
        try:
            with engine.begin() as conn:
                conn.execute(text(cmd))
        except Exception:
            pass

    try:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE staff_master ADD COLUMN staff_id TEXT;"))
    except Exception:
        pass

    try:
        with engine.begin() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM staff_master")).scalar()
            if result == 0:
                conn.execute(text("""
                    INSERT INTO staff_master (staff_name, password, role_name, is_admin, staff_id) 
                    VALUES ('店長', 'admin1234', '全体統括', TRUE, '0000')
                """))
    except Exception:
        pass

init_db()

# DB操作関数群
def load_staff():
    return pd.read_sql("SELECT * FROM staff_master", engine)

def load_day_data(day_str):
    return pd.read_sql(f"SELECT * FROM shift_data WHERE day = '{day_str}'", engine)

def get_all_shift_data():
    return pd.read_sql("SELECT * FROM shift_data", engine)

def save_day_data(day_str, df):
    with engine.begin() as conn:
        conn.execute(text(f"DELETE FROM shift_data WHERE day = '{day_str}'"))
        for _, row in df.iterrows():
            if row["氏名"] == "合計ライン": continue 
            # タイムスロットのデータをカンマ区切りに変換
            shift_values = ",".join([str(row.get(t, "")) for t in time_slots])
            conn.execute(text("""
                INSERT INTO shift_data (day, staff_name, off_status, shift_json)
                VALUES (:day, :staff_name, :off_status, :shift_json)
            """), {"day": day_str, "staff_name": row["氏名"], "off_status": row["休み"], "shift_json": shift_values})

def get_config(key, default_value=""):
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SELECT config_value FROM system_config WHERE config_key = :key"), {"key": key}).scalar()
            return result if result else default_value
    except:
        return default_value

def save_config(key, value):
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO system_config (config_key, config_value) 
            VALUES (:key, :value) 
            ON CONFLICT (config_key) DO UPDATE SET config_value = :value
        """), {"key": key, "value": str(value)})

# ==========================================
# 2. セッション管理 & 固定値
# ==========================================
if "user" not in st.session_state:
    st.session_state.user = None

time_slots = [f"{h}:{m:02d}" for h in range(8, 23) for m in (0, 30)]

if st.session_state.user is None:
    st.title("🔐 ログイン")
    with st.form("login_form"):
        input_name = st.text_input("氏名")
        input_pass = st.text_input("パスワード", type="password")
        if st.form_submit_button("ログイン", type="primary"):
            staff_df = load_staff()
            user_row = staff_df[(staff_df['staff_name'] == input_name) & (staff_df['password'] == input_pass)]
            if not user_row.empty:
                st.session_state.user = {"name": user_row.iloc[0]['staff_name'], "is_admin": user_row.iloc[0]['is_admin'], "role": user_row.iloc[0]['role_name']}
                st.rerun()
            else:
                st.error("認証失敗")
    st.stop()

st.sidebar.write(f"👤 {st.session_state.user['name']}")
if st.sidebar.button("ログアウト"):
    st.session_state.user = None
    st.rerun()

staff_df = load_staff()
staff_list = staff_df['staff_name'].tolist()
staff_id_map = dict(zip(staff_df['staff_name'], staff_df['staff_id']))

# ==========================================
# 3. 👨‍💼 管理者ダッシュボード
# ==========================================
if st.session_state.user["is_admin"]:
    st.title("👨‍💼 管理者メニュー")
    tab1, tab2, tab3, tab4 = st.tabs(["📝 シフト編集", "📅 募集設定", "👥 スタッフ管理", "📊 Excel出力"])
    
    with tab1:
        target_date = st.date_input("日付選択", value=datetime.date.today())
        target_day_str = target_date.strftime("%Y-%m-%d")
        
        # 週勤務計算
        start_of_week = target_date - datetime.timedelta(days=target_date.weekday())
        pd_start, pd_end = pd.to_datetime(start_of_week), pd.to_datetime(start_of_week + datetime.timedelta(days=6))
        all_df = get_all_shift_data()
        all_df['date_obj'] = pd.to_datetime(all_df['day'], errors='coerce')
        week_df = all_df[(all_df['date_obj'] >= pd_start) & (all_df['date_obj'] <= pd_end)].dropna(subset=['date_obj'])
        
        display_data = []
        total_counts = {t: 0 for t in time_slots}
        for s in staff_list:
            match = load_day_data(target_day_str).pipe(lambda d: d[d["staff_name"] == s])
            
            # 週勤務時間の集計
            wk_h = week_df[week_df['staff_name'] == s]['shift_json'].str.split(',').explode().isin(['1','2','同']).sum() * 0.5
            
            if not match.empty:
                # 🚨「未提出」を排除し、空なら空のまま表示
                status = match.iloc[0]["off_status"]
                row = {"ID": staff_id_map.get(s, ""), "氏名": s, "週勤務時間": f"{wk_h:.1f}", "休み": status}
                slots = match.iloc[0]["shift_json"].split(",")
                for j, t in enumerate(time_slots):
                    val = slots[j] if j < len(slots) else ""
                    row[t] = val
                    if val in ['1', '2', '同']: total_counts[t] += 1
            else:
                row = {"ID": staff_id_map.get(s, ""), "氏名": s, "週勤務時間": f"{wk_h:.1f}", "休み": "", **{t: "" for t in time_slots}}
            display_data.append(row)
        
        display_data.append({"ID": "", "氏名": "合計ライン", "週勤務時間": "-", "休み": "", **{t: str(total_counts[t]) for t in time_slots}})
        df_to_edit = pd.DataFrame(display_data)

        # AgGrid設定
        editable_js = JsCode("function(params) { return params.node.data['氏名'] !== '合計ライン'; }")
        cell_style_js = JsCode("""
        function(params) {
            const v = params.value;
            let style = {'fontSize': '11px', 'textAlign': 'center', 'padding': '0px', 'borderRight': '1px solid #e2e2e2', 'textOverflow': 'clip', 'whiteSpace': 'nowrap'};
            if (params.colDef.field === '氏名') style['textAlign'] = 'left';
            if (v === 'OFF' || v === '休') return Object.assign(style, {'backgroundColor': '#ff0000', 'color': '#ffffff'});
            if (v === '1') return Object.assign(style, {'backgroundColor': '#fce4d6'});
            if (v === '2') return Object.assign(style, {'backgroundColor': '#ffff00'});
            if (v === '同') return Object.assign(style, {'backgroundColor': '#00b050', 'color': '#ffffff'});
            if (params.node.data['氏名'] === '合計ライン') return Object.assign(style, {'backgroundColor': '#ffff00', 'fontWeight': 'bold'});
            return style;
        }
        """)

        work_calc_js = JsCode("function(params) { if (params.node.data['氏名'] === '合計ライン') return '-'; let active = 0; const ts = [" + ",".join([f"'{t}'" for t in time_slots]) + "]; ts.forEach(t => { if (['1', '2', '同'].includes(params.data[t])) active++; }); return active === 0 ? '0' : (active * 0.5).toFixed(1); }")
        break_calc_js = JsCode("function(params) { if (params.node.data['氏名'] === '合計ライン') return '-'; let brk = 0; const ts = [" + ",".join([f"'{t}'" for t in time_slots]) + "]; ts.forEach(t => { if (params.data[t] === '休') brk++; }); return brk === 0 ? '0' : (brk * 0.5).toFixed(1); }")

        # 列定義
        left_cols = [
            {"field": "ID", "pinned": "left", "width": 45, "editable": False, "cellStyle": cell_style_js},
            {"field": "氏名", "pinned": "left", "width": 85, "editable": False, "cellStyle": cell_style_js},
            {"headerName": "勤務h", "field": "勤務h", "pinned": "left", "width": 45, "editable": False, "valueGetter": work_calc_js, "cellStyle": cell_style_js},
            {"headerName": "休憩h", "field": "休憩h", "pinned": "left", "width": 45, "editable": False, "valueGetter": break_calc_js, "cellStyle": cell_style_js},
            # 🚨選択肢から「未提出」を削除
            {"field": "休み", "pinned": "left", "width": 50, "editable": editable_js, "cellEditor": 'agSelectCellEditor', "cellEditorParams": {'values': ["", "OFF"]}, "cellStyle": cell_style_js}
        ]
        time_cols = [{"headerName": f"{h}", "children": [{"field": f"{h}:{m:02d}", "headerName": "", "width": 25, "editable": editable_js, "cellEditor": 'agSelectCellEditor', "cellEditorParams": {'values': ["", "1", "2", "同", "休"]}, "cellStyle": cell_style_js} for m in (0, 30)]} for h in range(8, 23)]
        right_cols = [{"field": "週勤務時間", "pinned": "right", "width": 60, "editable": False, "cellStyle": cell_style_js}]

        gb = GridOptionsBuilder.from_dataframe(df_to_edit)
        grid_options = {"columnDefs": left_cols + time_cols + right_cols, "defaultColDef": {"sortable": False, "suppressMenu": True, "resizable": True, "suppressMovable": True}, "enableRangeSelection": True, "suppressCopyRowsToClipboard": True, "enterMovesDownAfterEdit": True, "singleClickEdit": True, "rowHeight": 22, "headerHeight": 22, "groupHeaderHeight": 22}
        
        response = AgGrid(df_to_edit, gridOptions=grid_options, update_mode=GridUpdateMode.VALUE_CHANGED, allow_unsafe_jscode=True, theme='balham', height=500)
        
        # オートセーブ
        edited_df = pd.DataFrame(response['data'])
        if not edited_df.empty and not df_to_edit.empty:
            if not edited_df.equals(df_to_edit):
                save_day_data(target_day_str, edited_df)
                st.toast("✅ 自動保存完了")
                st.rerun()

    # タブ2〜4（前回のImage 2再現ロジック含む）はそのまま維持
    # タブ2,3,4 は変更なしのため省略

    # 【タブ2】募集設定
    with tab2:
        st.write("従業員画面に表示するシフト提出の対象期間を設定します。")
        col1, col2 = st.columns(2)
        with col1:
            period = st.date_input("📅 シフト対象期間", value=(today, today + datetime.timedelta(days=6)))
        with col2:
            deadline = st.date_input("⏳ 提出締め切り日", value=today + datetime.timedelta(days=3))
            
        if st.button("募集設定を更新", type="primary"):
            if isinstance(period, tuple) and len(period) == 2:
                save_config("period_start", period[0].strftime("%Y-%m-%d"))
                save_config("period_end", period[1].strftime("%Y-%m-%d"))
                save_config("deadline", deadline.strftime("%Y-%m-%d"))
                st.success("✅ 募集設定を更新しました！")

    # 【タブ3】スタッフ管理（ID列追加）
    with tab3:
        st.write("#### 📁 CSVで一括インポート")
        @st.cache_data
        def get_csv_template():
            df_template = pd.DataFrame(columns=["ID", "氏名", "パスワード", "担当", "管理者権限"])
            df_template.loc[0] = ["1001", "テスト太郎", "pass123", "ホール", "FALSE"]
            return df_template.to_csv(index=False).encode('utf-8-sig')
            
        st.download_button("📥 インポート用CSVテンプレート", data=get_csv_template(), file_name="staff_template.csv", mime="text/csv")
        uploaded_file = st.file_uploader("CSVファイルをアップロード", type=["csv"])
        
        if uploaded_file:
            try:
                df_csv = pd.read_csv(uploaded_file)
                if list(df_csv.columns)[:5] != ["ID", "氏名", "パスワード", "担当", "管理者権限"]:
                    st.error("⚠️ 列の定義が異なります。")
                else:
                    if st.button("CSVデータをデータベースに保存", type="primary"):
                        with engine.connect() as conn:
                            for _, row in df_csv.iterrows():
                                is_admin_val = str(row["管理者権限"]).strip().lower() in ['true', '1', 'yes', 'はい']
                                conn.execute(text("""
                                    INSERT INTO staff_master (staff_id, staff_name, password, role_name, is_admin)
                                    VALUES (:id, :name, :pass, :role, :is_admin)
                                    ON CONFLICT (staff_name) DO UPDATE 
                                    SET staff_id = :id, password = :pass, role_name = :role, is_admin = :is_admin
                                """), {"id": str(row["ID"]), "name": str(row["氏名"]), "pass": str(row["パスワード"]), "role": str(row["担当"]), "is_admin": is_admin_val})
                            conn.commit()
                        st.success("✅ インポート完了！")
                        st.rerun()
            except Exception as e:
                st.error(f"エラー: {e}")
        
        st.write("---")
        st.write("#### ✍️ 手動編集")
        edited_staff = st.data_editor(
            staff_df, 
            column_config={
                "staff_id": st.column_config.TextColumn("ID"),
                "staff_name": st.column_config.TextColumn("氏名"),
                "password": st.column_config.TextColumn("パスワード"),
                "role_name": st.column_config.TextColumn("担当"),
                "is_admin": st.column_config.CheckboxColumn("管理者権限")
            },
            num_rows="dynamic", hide_index=True)
        
        if st.button("手動編集を保存"):
            with engine.connect() as conn:
                conn.execute(text("DELETE FROM staff_master")) 
                for _, row in edited_staff[edited_staff['staff_name'].str.strip() != ""].iterrows():
                    conn.execute(text("""
                        INSERT INTO staff_master (staff_id, staff_name, password, role_name, is_admin)
                        VALUES (:id, :name, :pass, :role, :is_admin)
                    """), {"id": str(row.get("staff_id", "")), "name": row["staff_name"], "pass": row["password"], "role": row["role_name"], "is_admin": row["is_admin"]})
                conn.commit()
            st.success("✅ 更新完了！")
            st.rerun()

    # 【タブ4】Excel出力（Image 2のタイル型完全再現）
    with tab4:
        st.write("### 📊 Excel出力")
        output_type = st.radio("出力形式を選択してください", ["日別出力（1日分）", "週別出力（7日分のタイル状配置）"])
        ex_start = st.date_input("開始日（週別の場合はこの日から7日間）", value=today)
        
        def write_day_block(ws, start_row, start_col, day_date):
            day_str = day_date.strftime("%Y-%m-%d")
            display_day = day_date.strftime("%m/%d")
            
            # ヘッダー構築
            headers = ["ID", "氏名", "勤務h", "休憩h", "休み"] + time_slots + ["週勤務"]
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=start_row, column=start_col + col_idx)
                # 時間の場合は「08:00」を「8」のように簡易表示（Image 2用）
                if ":" in header:
                    cell.value = ""
                else:
                    cell.value = header
                cell.font = Font(size=9, bold=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 日付ラベルを左上に
            ws.cell(row=start_row-1, column=start_col).value = display_day
            ws.cell(row=start_row-1, column=start_col).font = Font(bold=True)

            # データ取得
            df = load_day_data(day_str)
            all_df = get_all_shift_data()
            wk_start = day_date - datetime.timedelta(days=day_date.weekday())
            wk_end = wk_start + datetime.timedelta(days=6)
            all_df['date_obj'] = pd.to_datetime(all_df['day'], errors='coerce')
            week_df = all_df[(all_df['date_obj'] >= pd.to_datetime(wk_start)) & (all_df['date_obj'] <= pd.to_datetime(wk_end))]
            
            current_row = start_row + 1
            for s in staff_list:
                s_id = staff_id_map.get(s, "")
                match = df[df["staff_name"] == s]
                
                # 週勤務計算
                staff_wk = week_df[week_df['staff_name'] == s]
                wk_h = 0
                for _, r in staff_wk.iterrows():
                    slots = r["shift_json"].split(",")
                    wk_h += sum([1 for slot in slots if slot in ['1', '2', '同']])
                wk_h_str = f"{wk_h * 0.5:.1f}"

                if not match.empty:
                    status = match.iloc[0]["off_status"]
                    shift_vals = match.iloc[0]["shift_json"].split(",")
                    
                    work_h = sum([1 for x in shift_vals if x in ['1','2','同']]) * 0.5
                    break_h = sum([1 for x in shift_vals if x == '休']) * 0.5
                    
                    row_data = [s_id, s, work_h if work_h>0 else "0", break_h if break_h>0 else "0", status]
                    for i in range(len(time_slots)):
                        row_data.append(shift_vals[i] if i < len(shift_vals) else "")
                    row_data.append(wk_h_str)
                else:
                    row_data = [s_id, s, "0", "0", "未提出"] + [""] * len(time_slots) + [wk_h_str]
                
                # Excelへ書き込み＆色付け
                for col_idx, val in enumerate(row_data):
                    cell = ws.cell(row=current_row, column=start_col + col_idx)
                    cell.value = val
                    cell.font = Font(size=9)
                    cell.border = Border(left=Side(style='dotted', color='CCCCCC'), right=Side(style='dotted', color='CCCCCC'), bottom=Side(style='thin', color='EEEEEE'))
                    cell.alignment = Alignment(horizontal='center')
                    
                    # 色の適用（Image 1/2 仕様）
                    if val == '1': cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    elif val == '2': cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    elif val in ['休', 'OFF']: 
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(size=9, color="FFFFFF")
                    elif val == '同': 
                        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                        cell.font = Font(size=9, color="FFFFFF")
                        
                current_row += 1
            
            # 合計ライン
            ws.cell(row=current_row, column=start_col+1).value = "合計ライン"
            ws.cell(row=current_row, column=start_col+1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            return current_row + 2 # 次のブロック用の間隔

        def export_excel():
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "シフト表"
            
            # 列幅の調整（時間列は狭く）
            for c in range(1, 100): ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 3
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 12
            
            if output_type == "日別出力（1日分）":
                write_day_block(ws, 2, 1, ex_start)
                filename = f"シフト表_{ex_start.strftime('%Y%m%d')}.xlsx"
            else:
                # 週別出力（Image 2のタイル配置: 左列に月火水木、右列に金土日）
                # 簡略化のため、左列にDay1〜Day4、右列にDay5〜Day7を配置
                col_offset = len(time_slots) + 7
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_offset)].width = 5 # 列間の余白
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_offset+1)].width = 5
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_offset+2)].width = 12
                
                left_row = 2
                right_row = 2
                for i in range(7):
                    target_d = ex_start + datetime.timedelta(days=i)
                    if i < 4:
                        left_row = write_day_block(ws, left_row, 1, target_d)
                    else:
                        right_row = write_day_block(ws, right_row, col_offset + 1, target_d)
                        
                filename = f"週間シフト_{ex_start.strftime('%m%d')}-{(ex_start + datetime.timedelta(days=6)).strftime('%m%d')}.xlsx"
                
            wb.save(output)
            return output.getvalue(), filename

        import openpyxl # Excel処理用に追加
        excel_data, dl_filename = export_excel()
        st.download_button(label=f"📥 {output_type}をExcelでダウンロード", data=excel_data, file_name=dl_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

# ==========================================
# 4. 📱 従業員メニュー
# ==========================================
else:
    st.title("📱 シフト希望提出")
    p_start_str = get_config("period_start")
    p_end_str = get_config("period_end")
    p_deadline = get_config("deadline", "未設定")
    
    if p_start_str and p_end_str:
        st.warning(f"**現在の募集期間:** {p_start_str} 〜 {p_end_str} 　🚨 **提出締切:** {p_deadline}")
        p_start = datetime.datetime.strptime(p_start_str, "%Y-%m-%d").date()
        p_end = datetime.datetime.strptime(p_end_str, "%Y-%m-%d").date()
        target_dates = [(p_start + datetime.timedelta(days=i)).strftime("%Y-%m-%d") for i in range((p_end - p_start).days + 1)]
        
        off_days = st.multiselect("お休み（OFF）希望の日付を選んでください", target_dates)
        
        if st.button("シフトを提出する", type="primary"):
            my_name = st.session_state.user["name"]
            for d_str in target_dates:
                df = load_day_data(d_str)
                status = "OFF" if d_str in off_days else ""
                if my_name in df["staff_name"].values:
                    df.loc[df["staff_name"] == my_name, "off_status"] = status
                else:
                    new_row = pd.DataFrame([{"day": d_str, "staff_name": my_name, "off_status": status, "shift_json": ",".join([""]*len(time_slots))}])
                    df = pd.concat([df, new_row], ignore_index=True)
                
                save_df = pd.DataFrame([{"氏名": r["staff_name"], "休み": r["off_status"], **dict(zip(time_slots, r["shift_json"].split(",")))} for _, r in df.iterrows()])
                save_day_data(d_str, save_df)
                
            st.success("✅ あなたのシフト希望を提出しました！")
    else:
        st.error("現在、管理者によって設定されたシフト募集期間がありません。")
